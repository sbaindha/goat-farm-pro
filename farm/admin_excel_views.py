"""
Admin-Only Excel Download & Upload Views
फाइल: farm/admin_excel_views.py

केवल admin users को access मिलेगा
"""

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
import pandas as pd
from datetime import datetime
import traceback

from .models import (
    Goat, BreedingRecord, HealthRecord, MilkProduction,
    Sale, Expense, WeightRecord, Task, Customer, Credit, 
    Notification, Insurance, MortalityRecord, AdditionalIncome, 
    ActivityLog, VetVisit, VaccinationSchedule, BudgetPlanning, 
    PerformanceEvaluation, CustomReminder, Document, PhotoGallery, 
    WeatherRecord, MarketPrice, FarmEvent, BreedingPlan
)
from .excel_export import export_to_excel


# ==================== HELPER FUNCTION ====================

def is_admin(user):
    """Check if user is admin"""
    return user.is_staff or user.is_superuser


def admin_required(view_func):
    """Decorator: Admin-only access"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not is_admin(request.user):
            messages.error(request, '❌ केवल admin users को access है!')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ==================== EXCEL DOWNLOAD VIEWS (ADMIN ONLY) ====================

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_goats_excel(request):
    """Admin: बकरियों की सूची को Excel में download करो"""
    return export_to_excel('goats', request)

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_breeding_excel(request):
    """Admin: Breeding records को Excel में download करो"""
    return export_to_excel('breeding', request)

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_health_excel(request):
    """Admin: Health records को Excel में download करो"""
    return export_to_excel('health', request)

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_milk_excel(request):
    """Admin: Milk production को Excel में download करो"""
    return export_to_excel('milk', request)

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_sales_excel(request):
    """Admin: Sales को Excel में download करो"""
    return export_to_excel('sales', request)

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_expenses_excel(request):
    """Admin: Expenses को Excel में download करो"""
    return export_to_excel('expenses', request)

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_download_all_data(request):
    """Admin: सभी data को एक Excel file में download करो"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    wb.remove(wb.active)  # पहली blank sheet को remove करो
    
    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    
    # ==================== GOATS SHEET ====================
    ws = wb.create_sheet("Goats (बकरियां)")
    goats = Goat.objects.all()
    headers = ['Tag Number', 'Name', 'Breed', 'Gender', 'Color', 'DOB', 
              'Age (Years)', 'Weight (kg)', 'Purchase Date', 'Purchase Price (₹)', 
              'Status', 'Created At']
    ws.append(headers)
    
    for goat in goats:
        ws.append([
            goat.tag_number, goat.name, goat.breed, goat.gender, goat.color,
            goat.date_of_birth.strftime('%d-%m-%Y'), goat.get_age_years(), 
            goat.weight, goat.purchase_date.strftime('%d-%m-%Y'),
            goat.purchase_price, goat.get_status_display(), 
            goat.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== HEALTH RECORDS SHEET ====================
    ws = wb.create_sheet("Health (स्वास्थ्य)")
    health = HealthRecord.objects.select_related('goat')
    headers = ['Goat', 'Type', 'Date', 'Description', 'Medicine', 'Dosage', 
              'Cost (₹)', 'Created At']
    ws.append(headers)
    
    for record in health:
        ws.append([
            record.goat.name, record.get_record_type_display(),
            record.date.strftime('%d-%m-%Y'), record.description,
            record.medicine_used, record.dosage, record.cost,
            record.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== MILK PRODUCTION SHEET ====================
    ws = wb.create_sheet("Milk (दूध)")
    milk = MilkProduction.objects.select_related('goat')
    headers = ['Goat', 'Date', 'Session', 'Quantity (L)', 'Fat %', 'Created At']
    ws.append(headers)
    
    for record in milk:
        ws.append([
            record.goat.name, record.date.strftime('%d-%m-%Y'),
            record.session, record.quantity, record.fat_percentage or '',
            record.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== SALES SHEET ====================
    ws = wb.create_sheet("Sales (बिक्रय)")
    sales = Sale.objects.select_related('goat')
    headers = ['Type', 'Goat', 'Date', 'Quantity', 'Unit', 'Price per Unit (₹)', 
              'Total Amount (₹)', 'Buyer', 'Created At']
    ws.append(headers)
    
    for sale in sales:
        ws.append([
            sale.sale_type, sale.goat.name if sale.goat else '',
            sale.date.strftime('%d-%m-%Y'), sale.quantity, sale.unit,
            sale.price_per_unit, sale.total_amount, sale.buyer_name,
            sale.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== EXPENSES SHEET ====================
    ws = wb.create_sheet("Expenses (खर्च)")
    expenses = Expense.objects.all()
    headers = ['Date', 'Type', 'Description', 'Amount (₹)', 'Paid To', 'Created At']
    ws.append(headers)
    
    for expense in expenses:
        ws.append([
            expense.date.strftime('%d-%m-%Y'), expense.expense_type,
            expense.description, expense.amount, expense.paid_to,
            expense.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== SUMMARY SHEET ====================
    ws = wb.create_sheet("Summary (सारांश)", 0)
    
    from django.db.models import Sum
    
    total_milk = MilkProduction.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_revenue = Sale.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    
    ws['A1'] = "🐐 GOAT FARM MANAGEMENT - DATA SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = "Export Date"
    ws['B3'] = timestamp
    
    ws['A5'] = "STATISTICS"
    ws['A5'].font = Font(bold=True, size=12)
    
    ws['A6'] = "Total Goats"
    ws['B6'] = Goat.objects.count()
    
    ws['A7'] = "Active Goats"
    ws['B7'] = Goat.objects.filter(status='A').count()
    
    ws['A8'] = "Total Milk Production (L)"
    ws['B8'] = round(total_milk, 2)
    
    ws['A9'] = "Total Sales Revenue (₹)"
    ws['B9'] = round(total_revenue, 2)
    
    ws['A10'] = "Total Expenses (₹)"
    ws['B10'] = round(total_expenses, 2)
    
    ws['A11'] = "Net Profit (₹)"
    ws['B11'] = round(total_revenue - total_expenses, 2)
    
    ws['A11'].font = Font(bold=True)
    ws['B11'].font = Font(bold=True)
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="farm_complete_data_{datetime.now().strftime("%d-%m-%Y")}.xlsx"'
    wb.save(response)
    
    return response


# ==================== EXCEL IMPORT VIEWS (ADMIN ONLY) ====================

@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET", "POST"])
def admin_excel_import(request):
    """Admin: Excel से data import करो"""
    
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('file')
            model_name = request.POST.get('model')
            
            if not excel_file:
                messages.error(request, '❌ File select करो!')
                return redirect('admin_excel_page')
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '❌ केवल Excel files (.xlsx, .xls) support हैं!')
                return redirect('admin_excel_page')
            
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            imported_count = 0
            skipped_count = 0
            
            # Import based on model
            if model_name == 'goats':
                imported_count, skipped_count = import_goats_from_excel(df)
            elif model_name == 'health':
                imported_count, skipped_count = import_health_from_excel(df)
            elif model_name == 'milk':
                imported_count, skipped_count = import_milk_from_excel(df)
            elif model_name == 'sales':
                imported_count, skipped_count = import_sales_from_excel(df)
            elif model_name == 'expenses':
                imported_count, skipped_count = import_expenses_from_excel(df)
            
            messages.success(
                request, 
                f'✅ Import successful! {imported_count} records added, {skipped_count} skipped'
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
            print(traceback.format_exc())
        
        return redirect('admin_excel_page')
    
    # GET request - show import form
    return render(request, 'farm/admin_excel.html', {
        'page': 'import'
    })


@login_required(login_url='/login/')
@admin_required
@require_http_methods(["GET"])
def admin_excel_page(request):
    """Admin: Excel downloads/uploads का main page"""
    context = {
        'download_options': [
            {'name': 'बकरियां (Goats)', 'url': 'admin_download_goats', 'icon': '🐐'},
            {'name': 'प्रजनन रिकॉर्ड (Breeding)', 'url': 'admin_download_breeding', 'icon': '👨‍👩‍👧'},
            {'name': 'स्वास्थ्य रिकॉर्ड (Health)', 'url': 'admin_download_health', 'icon': '🏥'},
            {'name': 'दूध उत्पादन (Milk)', 'url': 'admin_download_milk', 'icon': '🥛'},
            {'name': 'बिक्रय (Sales)', 'url': 'admin_download_sales', 'icon': '💰'},
            {'name': 'खर्च (Expenses)', 'url': 'admin_download_expenses', 'icon': '💸'},
            {'name': 'पूरा डेटा (Complete Data)', 'url': 'admin_download_all', 'icon': '📦'},
        ],
        'import_options': [
            {'name': 'बकरियां (Goats)', 'model': 'goats', 'icon': '🐐'},
            {'name': 'स्वास्थ्य रिकॉर्ड (Health)', 'model': 'health', 'icon': '🏥'},
            {'name': 'दूध उत्पादन (Milk)', 'model': 'milk', 'icon': '🥛'},
            {'name': 'बिक्रय (Sales)', 'model': 'sales', 'icon': '💰'},
            {'name': 'खर्च (Expenses)', 'model': 'expenses', 'icon': '💸'},
        ]
    }
    return render(request, 'farm/admin_excel.html', context)


# ==================== HELPER FUNCTIONS FOR IMPORT ====================

def import_goats_from_excel(df):
    """Excel से Goats import करो"""
    imported = 0
    skipped = 0
    
    for idx, row in df.iterrows():
        try:
            tag_number = str(row.get('Tag Number', '')).strip()
            
            if not tag_number or pd.isna(tag_number):
                skipped += 1
                continue
            
            # Check if already exists
            if Goat.objects.filter(tag_number=tag_number).exists():
                skipped += 1
                continue
            
            # Create goat
            Goat.objects.create(
                tag_number=tag_number,
                name=str(row.get('Name', tag_number)),
                breed=str(row.get('Breed', 'local')).lower(),
                gender=str(row.get('Gender', 'M')[0]).upper(),
                color=str(row.get('Color', 'Brown')),
                date_of_birth=pd.to_datetime(row.get('DOB')),
                weight=float(row.get('Weight (kg)', 0)) or 0,
                purchase_date=pd.to_datetime(row.get('Purchase Date')),
                purchase_price=float(row.get('Purchase Price (₹)', 0)) or 0,
                status=str(row.get('Status', 'A')[0]).upper(),
            )
            imported += 1
        except Exception as e:
            skipped += 1
            print(f"Error importing row {idx}: {str(e)}")
    
    return imported, skipped


def import_health_from_excel(df):
    """Excel से Health records import करो"""
    imported = 0
    skipped = 0
    
    for idx, row in df.iterrows():
        try:
            goat_name = str(row.get('Goat', '')).strip()
            
            try:
                goat = Goat.objects.get(name=goat_name)
            except Goat.DoesNotExist:
                skipped += 1
                continue
            
            HealthRecord.objects.create(
                goat=goat,
                record_type=str(row.get('Type', 'C')[0]).upper(),
                date=pd.to_datetime(row.get('Date')),
                description=str(row.get('Description', '')),
                medicine_used=str(row.get('Medicine', '')),
                dosage=str(row.get('Dosage', '')),
                cost=float(row.get('Cost (₹)', 0)) or 0,
                veterinarian=str(row.get('Veterinarian', '')),
            )
            imported += 1
        except Exception as e:
            skipped += 1
            print(f"Error importing row {idx}: {str(e)}")
    
    return imported, skipped


def import_milk_from_excel(df):
    """Excel से Milk production import करो"""
    imported = 0
    skipped = 0
    
    for idx, row in df.iterrows():
        try:
            goat_name = str(row.get('Goat', '')).strip()
            
            try:
                goat = Goat.objects.get(name=goat_name)
            except Goat.DoesNotExist:
                skipped += 1
                continue
            
            MilkProduction.objects.create(
                goat=goat,
                date=pd.to_datetime(row.get('Date')),
                session=str(row.get('Session', 'M')),
                quantity=float(row.get('Quantity (L)', 0)) or 0,
                fat_percentage=float(row.get('Fat %', 0)) if pd.notna(row.get('Fat %')) else None,
            )
            imported += 1
        except Exception as e:
            skipped += 1
            print(f"Error importing row {idx}: {str(e)}")
    
    return imported, skipped


def import_sales_from_excel(df):
    """Excel से Sales import करो"""
    imported = 0
    skipped = 0
    
    for idx, row in df.iterrows():
        try:
            goat_name = str(row.get('Goat', '')).strip()
            goat = None
            
            if goat_name:
                try:
                    goat = Goat.objects.get(name=goat_name)
                except Goat.DoesNotExist:
                    pass
            
            Sale.objects.create(
                sale_type=str(row.get('Type', 'milk')),
                goat=goat,
                date=pd.to_datetime(row.get('Date')),
                quantity=float(row.get('Quantity', 0)) or 0,
                unit=str(row.get('Unit', 'L')),
                price_per_unit=float(row.get('Price per Unit (₹)', 0)) or 0,
                total_amount=float(row.get('Total Amount (₹)', 0)) or 0,
                buyer_name=str(row.get('Buyer', 'Unknown')),
                buyer_contact=str(row.get('Contact', '')),
            )
            imported += 1
        except Exception as e:
            skipped += 1
            print(f"Error importing row {idx}: {str(e)}")
    
    return imported, skipped


def import_expenses_from_excel(df):
    """Excel से Expenses import करो"""
    imported = 0
    skipped = 0
    
    for idx, row in df.iterrows():
        try:
            Expense.objects.create(
                date=pd.to_datetime(row.get('Date')),
                expense_type=str(row.get('Type', 'feed')),
                description=str(row.get('Description', '')),
                amount=float(row.get('Amount (₹)', 0)) or 0,
                paid_to=str(row.get('Paid To', 'Unknown')),
                payment_method=str(row.get('Payment Method', 'C')[0]).upper() if pd.notna(row.get('Payment Method')) else 'C',
            )
            imported += 1
        except Exception as e:
            skipped += 1
            print(f"Error importing row {idx}: {str(e)}")
    
    return imported, skipped
