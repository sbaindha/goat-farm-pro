from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from datetime import datetime, timedelta

# Use specific imports instead of wildcard
from .models import (
    Goat, BreedingRecord, HealthRecord, MilkProduction,
    FeedInventory, FeedConsumption, Sale, Expense, WeightRecord,
    Task, Customer, Credit, Notification, Insurance, MortalityRecord,
    AdditionalIncome, ActivityLog, VetVisit, VaccinationSchedule,
    BudgetPlanning, PerformanceEvaluation, CustomReminder, Document,
    PhotoGallery, WeatherRecord, MarketPrice, FarmEvent, BreedingPlan
)



from django.http import HttpResponse

from .excel_export import export_to_excel


def login_view(request):
    """Login Page"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    error = ''
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            next_url = request.GET.get('next', '/')
            return redirect(next_url)
        else:
            error = 'Invalid username ya password. Dobara try karein.'
    return render(request, 'farm/login.html', {'error': error})


def logout_view(request):
    """Logout"""
    logout(request)
    return redirect('login')


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def dashboard(request):
    """Professional Dashboard - मुख्य डैशबोर्ड"""
    from datetime import date
    today = date.today()
    context = {
        'total_goats':        Goat.objects.count(),
        'active_goats':       Goat.objects.filter(status='A').count(),
        'pregnant_goats':     Goat.objects.filter(status='P').count(),
        'male_goats':         Goat.objects.filter(gender='M').count(),
        'female_goats':       Goat.objects.filter(gender='F').count(),
        'recent_goats':       Goat.objects.order_by('-created_at')[:6],
        'recent_health':      HealthRecord.objects.select_related('goat').order_by('-date')[:5],
        'upcoming_deliveries':BreedingRecord.objects.select_related('mother','father').filter(
                                  status__in=['P','C'], expected_delivery_date__gte=today
                              ).order_by('expected_delivery_date')[:4],
        'total_milk_today':   MilkProduction.objects.filter(date=today).aggregate(
                                  total=Sum('quantity'))['total'] or 0,
        'total_sales_amount': Sale.objects.aggregate(total=Sum('total_amount'))['total'] or 0,
        'total_expenses_amount': Expense.objects.aggregate(total=Sum('amount'))['total'] or 0,
        'pending_tasks':      Task.objects.filter(status='P').count(),
        'overdue_vaccinations': VaccinationSchedule.objects.filter(
                                  completed=False, due_date__lt=today).count(),
        'pending_reminders':  CustomReminder.objects.filter(is_active=True).count(),
        'recent_sales':       Sale.objects.order_by('-date')[:5],
    }
    context['profit'] = context['total_sales_amount'] - context['total_expenses_amount']
    return render(request, 'farm/professional_dashboard.html', context)

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def weather_dashboard(request):
    """Weather Dashboard - मौसम डैशबोर्ड"""
    return render(request, 'farm/weather_dashboard.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def goats_list(request):
    """Goats Management - बकरियां प्रबंधन"""
    return render(request, 'farm/goats_management.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def breeding_list(request):
    """Breeding Records - प्रजनन रिकॉर्ड"""
    return render(request, 'farm/breeding_management.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def health_list(request):
    """Health Records - स्वास्थ्य रिकॉर्ड"""
    return render(request, 'farm/health_management.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def milk_list(request):
    """Milk Production - दूध उत्पादन"""
    return render(request, 'farm/milk_management.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def sales_list(request):
    """Sales Management - बिक्रय प्रबंधन"""
    return render(request, 'farm/sales_management.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def expenses_list(request):
    """Expenses Management - खर्च प्रबंधन"""
    return render(request, 'farm/expenses_management.html')

# FIX #1: Removed duplicate @login_required decorator (was on lines 82-83)
@login_required(login_url='/login/')
@require_http_methods(["GET"])
def api_docs(request):
    """Purana URL — naye API Help page par redirect karo."""
    from django.shortcuts import redirect
    return redirect('api_help')


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def api_help(request):
    """Interactive API Help page — har endpoint ki help + live try."""
    return render(request, 'farm/api_help.html')


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_all_data(request):
    """
    सभी data को एक Excel file में download करो
    Multiple sheets के साथ
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    wb.remove(wb.active)  # पहली blank sheet को remove करो
    
    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    
    # ==================== GOATS SHEET ====================
    ws = wb.create_sheet("Goats (बकरियां)")
    goats = Goat.objects.all()
    headers = ['Tag Number', 'Name', 'Breed', 'Gender', 'Color', 'DOB', 
              'Age (Years)', 'Weight (kg)', 'Purchase Date', 'Purchase Price (₹)', 
              'Status', 'Created At']
    ws.append(headers)
    
    for goat in goats:
        ws.append([
            goat.tag_number, goat.name, goat.breed, goat.gender, goat.color,
            goat.date_of_birth.strftime('%d-%m-%Y'), goat.get_age_years(), 
            goat.weight, goat.purchase_date.strftime('%d-%m-%Y'),
            goat.purchase_price, goat.get_status_display(), 
            goat.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== HEALTH RECORDS SHEET ====================
    ws = wb.create_sheet("Health (स्वास्थ्य)")
    health = HealthRecord.objects.select_related('goat')
    headers = ['Goat', 'Type', 'Date', 'Description', 'Medicine', 'Dosage', 
              'Cost (₹)', 'Created At']
    ws.append(headers)
    
    for record in health:
        ws.append([
            record.goat.name, record.get_record_type_display(),
            record.date.strftime('%d-%m-%Y'), record.description,
            record.medicine_used, record.dosage, record.cost,
            record.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== MILK PRODUCTION SHEET ====================
    ws = wb.create_sheet("Milk (दूध)")
    milk = MilkProduction.objects.select_related('goat')
    headers = ['Goat', 'Date', 'Session', 'Quantity (L)', 'Fat %', 'Created At']
    ws.append(headers)
    
    for record in milk:
        ws.append([
            record.goat.name, record.date.strftime('%d-%m-%Y'),
            record.session, record.quantity, record.fat_percentage or '',
            record.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== SALES SHEET ====================
    ws = wb.create_sheet("Sales (बिक्रय)")
    sales = Sale.objects.select_related('goat')
    headers = ['Type', 'Goat', 'Date', 'Quantity', 'Unit', 'Price per Unit (₹)', 
              'Total Amount (₹)', 'Buyer', 'Created At']
    ws.append(headers)
    
    for sale in sales:
        ws.append([
            sale.sale_type, sale.goat.name if sale.goat else '',
            sale.date.strftime('%d-%m-%Y'), sale.quantity, sale.unit,
            sale.price_per_unit, sale.total_amount, sale.buyer_name,
            sale.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== EXPENSES SHEET ====================
    ws = wb.create_sheet("Expenses (खर्च)")
    expenses = Expense.objects.all()
    headers = ['Date', 'Type', 'Description', 'Amount (₹)', 'Paid To', 'Created At']
    ws.append(headers)
    
    for expense in expenses:
        ws.append([
            expense.date.strftime('%d-%m-%Y'), expense.expense_type,
            expense.description, expense.amount, expense.paid_to,
            expense.created_at.strftime('%d-%m-%Y %H:%M')
        ])
    
    # ==================== SUMMARY SHEET ====================
    ws = wb.create_sheet("Summary (सारांश)", 0)
    
    total_milk = MilkProduction.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_revenue = Sale.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    
    ws['A1'] = "🐐 GOAT FARM MANAGEMENT - DATA SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = "Export Date"
    ws['B3'] = timestamp
    
    ws['A5'] = "STATISTICS"
    ws['A5'].font = Font(bold=True, size=12)
    
    ws['A6'] = "Total Goats"
    ws['B6'] = Goat.objects.count()
    
    ws['A7'] = "Active Goats"
    ws['B7'] = Goat.objects.filter(status='A').count()
    
    ws['A8'] = "Total Milk Production (L)"
    ws['B8'] = round(total_milk, 2)
    
    ws['A9'] = "Total Sales Revenue (₹)"
    ws['B9'] = round(total_revenue, 2)
    
    ws['A10'] = "Total Expenses (₹)"
    ws['B10'] = round(total_expenses, 2)
    
    ws['A11'] = "Net Profit (₹)"
    ws['B11'] = round(total_revenue - total_expenses, 2)
    
    ws['A11'].font = Font(bold=True)
    ws['B11'].font = Font(bold=True)
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="farm_complete_data_{datetime.now().strftime("%d-%m-%Y")}.xlsx"'
    wb.save(response)
    
    return response


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def export_downloads_page(request):
    """Excel downloads aur Backup ka page — ab /backup/ par redirect karo"""
    from django.shortcuts import redirect
    return redirect('backup')


# ==================== EXCEL DOWNLOAD VIEWS ====================
# urls.py mein from .views import (...) ke liye required functions

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_goats_excel(request):
    """Goats data Excel mein download karo"""
    return export_to_excel('goats', request)

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_breeding_excel(request):
    """Breeding records Excel mein download karo"""
    return export_to_excel('breeding', request)

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_health_excel(request):
    """Health records Excel mein download karo"""
    return export_to_excel('health', request)

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_milk_excel(request):
    """Milk production Excel mein download karo"""
    return export_to_excel('milk', request)

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_sales_excel(request):
    """Sales data Excel mein download karo"""
    return export_to_excel('sales', request)

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_expenses_excel(request):
    """Expenses data Excel mein download karo"""
    return export_to_excel('expenses', request)


# ==================== v6.0 NEW VIEWS ====================

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def analytics_dashboard(request):
    """Analytics Dashboard — Charts, P&L, Breed Performance"""
    return render(request, 'farm/analytics_dashboard.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def ai_dashboard(request):
    """AI Insights Dashboard — Breeding, Sick Detection, Sell Suggestions"""
    return render(request, 'farm/ai_dashboard.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def invoice_management(request):
    """Invoice Management — All sales with PDF download"""
    return render(request, 'farm/invoice_management.html')

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def qr_management(request):
    """QR Code Management — Per-goat QR codes"""
    return render(request, 'farm/qr_management.html')

def goat_scan_page(request, tag_number):
    """
    Public page — QR scan se khulta hai.
    Mobile-friendly goat profile.
    Login required nahi — anyone with QR can view basic info.
    """
    from django.http import JsonResponse
    try:
        goat = Goat.objects.get(tag_number=tag_number)
        context = {
            'goat': goat,
            'recent_health': HealthRecord.objects.filter(goat=goat).order_by('-date')[:3],
            'latest_weight': WeightRecord.objects.filter(goat=goat).order_by('-date').first(),
            'total_milk': MilkProduction.objects.filter(goat=goat).aggregate(total=Sum('quantity'))['total'] or 0,
        }
        return render(request, 'farm/goat_scan.html', context)
    except Goat.DoesNotExist:
        from django.http import Http404
        raise Http404(f"Goat with tag '{tag_number}' not found")
