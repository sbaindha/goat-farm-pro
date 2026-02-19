"""
Excel Export Utility - सब models को Excel में export करो
फाइल: farm/excel_export.py
"""

import pandas as pd
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Goat, BreedingRecord, HealthRecord, MilkProduction,
    Sale, Expense, WeightRecord, PerformanceEvaluation,
    MarketPrice, WeatherRecord, FarmEvent, BreedingPlan,
    CustomReminder, AdditionalIncome, Task, Customer,
    Credit, Notification, Insurance, MortalityRecord,
    VaccinationSchedule, BudgetPlanning, ActivityLog, VetVisit
)


class ExcelExporter:
    """Excel में data export करने के लिए utility class"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Data"
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def style_headers(self, row_num, col_count):
        """Header को style करो"""
        for col in range(1, col_count + 1):
            cell = self.ws.cell(row=row_num, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def style_cells(self, start_row, end_row, col_count):
        """सभी cells को style करो"""
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def auto_adjust_columns(self):
        """Column width को auto adjust करो"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_goats(self):
        """सभी बकरियों को export करो"""
        goats = Goat.objects.all()
        
        headers = ['Tag Number', 'Name', 'Breed', 'Gender', 'Color', 'DOB', 
                  'Age (Years)', 'Weight (kg)', 'Purchase Date', 'Purchase Price (₹)', 
                  'Status', 'Mother', 'Father', 'Created At']
        
        self.ws.append(headers)
        self.style_headers(1, len(headers))
        
        row_num = 2
        for goat in goats:
            self.ws.append([
                goat.tag_number,
                goat.name,
                dict(Goat.BREED_CHOICES).get(goat.breed, goat.breed),
                dict(Goat.GENDER_CHOICES).get(goat.gender, goat.gender),
                goat.color,
                goat.date_of_birth.strftime('%d-%m-%Y'),
                goat.get_age_years(),
                goat.weight,
                goat.purchase_date.strftime('%d-%m-%Y'),
                goat.purchase_price,
                dict(Goat.STATUS_CHOICES).get(goat.status, goat.status),
                goat.mother.name if goat.mother else '',
                goat.father.name if goat.father else '',
                goat.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
            row_num += 1
        
        self.style_cells(2, row_num - 1, len(headers))
        self.auto_adjust_columns()
        return self.wb
    
    def export_breeding(self):
        """Breeding records export करो"""
        records = BreedingRecord.objects.select_related('mother', 'father')
        
        headers = ['Mother', 'Father', 'Breeding Date', 'Expected Delivery', 
                  'Actual Delivery', 'Status', 'Number of Kids', 'Notes', 'Created At']
        
        self.ws.append(headers)
        self.style_headers(1, len(headers))
        
        row_num = 2
        for record in records:
            self.ws.append([
                record.mother.name,
                record.father.name,
                record.breeding_date.strftime('%d-%m-%Y'),
                record.expected_delivery_date.strftime('%d-%m-%Y'),
                record.actual_delivery_date.strftime('%d-%m-%Y') if record.actual_delivery_date else '',
                dict(BreedingRecord.STATUS_CHOICES).get(record.status, record.status),
                record.number_of_kids or '',
                record.notes,
                record.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
            row_num += 1
        
        self.style_cells(2, row_num - 1, len(headers))
        self.auto_adjust_columns()
        return self.wb
    
    def export_health(self):
        """Health records export करो"""
        records = HealthRecord.objects.select_related('goat')
        
        headers = ['Goat', 'Type', 'Date', 'Description', 'Medicine', 'Dosage', 
                  'Cost (₹)', 'Veterinarian', 'Next Due Date', 'Created At']
        
        self.ws.append(headers)
        self.style_headers(1, len(headers))
        
        row_num = 2
        for record in records:
            self.ws.append([
                record.goat.name,
                dict(HealthRecord.RECORD_TYPE_CHOICES).get(record.record_type, record.record_type),
                record.date.strftime('%d-%m-%Y'),
                record.description,
                record.medicine_used,
                record.dosage,
                record.cost,
                record.veterinarian,
                record.next_due_date.strftime('%d-%m-%Y') if record.next_due_date else '',
                record.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
            row_num += 1
        
        self.style_cells(2, row_num - 1, len(headers))
        self.auto_adjust_columns()
        return self.wb
    
    def export_milk(self):
        """Milk production export करो"""
        records = MilkProduction.objects.select_related('goat')
        
        headers = ['Goat', 'Date', 'Session', 'Quantity (L)', 'Fat %', 'Created At']
        
        self.ws.append(headers)
        self.style_headers(1, len(headers))
        
        row_num = 2
        for record in records:
            self.ws.append([
                record.goat.name,
                record.date.strftime('%d-%m-%Y'),
                record.session,
                record.quantity,
                record.fat_percentage or '',
                record.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
            row_num += 1
        
        self.style_cells(2, row_num - 1, len(headers))
        self.auto_adjust_columns()
        return self.wb
    
    def export_sales(self):
        """Sales export करो"""
        sales = Sale.objects.select_related('goat')
        
        headers = ['Type', 'Goat', 'Date', 'Quantity', 'Unit', 'Price per Unit (₹)', 
                  'Total Amount (₹)', 'Buyer', 'Contact', 'Payment Status', 'Created At']
        
        self.ws.append(headers)
        self.style_headers(1, len(headers))
        
        row_num = 2
        for sale in sales:
            self.ws.append([
                sale.sale_type,
                sale.goat.name if sale.goat else '',
                sale.date.strftime('%d-%m-%Y'),
                sale.quantity,
                sale.unit,
                sale.price_per_unit,
                sale.total_amount,
                sale.buyer_name,
                sale.buyer_contact,
                dict(Sale.PAYMENT_STATUS_CHOICES).get(sale.payment_status, sale.payment_status) if hasattr(Sale, 'PAYMENT_STATUS_CHOICES') else sale.payment_status,
                sale.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
            row_num += 1
        
        self.style_cells(2, row_num - 1, len(headers))
        self.auto_adjust_columns()
        return self.wb
    
    def export_expenses(self):
        """Expenses export करो"""
        expenses = Expense.objects.all()
        
        headers = ['Date', 'Type', 'Description', 'Amount (₹)', 'Paid To', 
                  'Payment Method', 'Created At']
        
        self.ws.append(headers)
        self.style_headers(1, len(headers))
        
        row_num = 2
        for expense in expenses:
            self.ws.append([
                expense.date.strftime('%d-%m-%Y'),
                expense.expense_type,
                expense.description,
                expense.amount,
                expense.paid_to,
                dict(Expense.PAYMENT_METHOD_CHOICES).get(expense.payment_method, expense.payment_method) if hasattr(Expense, 'PAYMENT_METHOD_CHOICES') else expense.payment_method,
                expense.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
            row_num += 1
        
        self.style_cells(2, row_num - 1, len(headers))
        self.auto_adjust_columns()
        return self.wb
    
    def get_response(self, filename):
        """HTTP Response के रूप में Excel file return करो"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        self.wb.save(response)
        return response


def export_to_excel(model_name, request=None):
    """
    किसी भी model को Excel में export करो
    
    Usage in views:
        return export_to_excel('goats')
        return export_to_excel('breeding')
        return export_to_excel('health')
        return export_to_excel('milk')
        return export_to_excel('sales')
        return export_to_excel('expenses')
    """
    exporter = ExcelExporter()
    timestamp = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
    
    if model_name.lower() == 'goats':
        exporter.export_goats()
        filename = f'goats_{timestamp}.xlsx'
    elif model_name.lower() == 'breeding':
        exporter.export_breeding()
        filename = f'breeding_records_{timestamp}.xlsx'
    elif model_name.lower() == 'health':
        exporter.export_health()
        filename = f'health_records_{timestamp}.xlsx'
    elif model_name.lower() == 'milk':
        exporter.export_milk()
        filename = f'milk_production_{timestamp}.xlsx'
    elif model_name.lower() == 'sales':
        exporter.export_sales()
        filename = f'sales_{timestamp}.xlsx'
    elif model_name.lower() == 'expenses':
        exporter.export_expenses()
        filename = f'expenses_{timestamp}.xlsx'
    else:
        return None
    
    return exporter.get_response(filename)
