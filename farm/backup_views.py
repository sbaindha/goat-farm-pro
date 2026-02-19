"""
Backup & Excel Download Views
farm/backup_views.py

Features:
  - JSON full backup  (sab models, single file, restore ready)
  - JSON restore      (backup se data wapas laao)
  - Excel download    (individual sheets + complete workbook)
  - Backup history    (last N backups server par store)
"""

import json
import os
import io
import zipfile
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_protect  # FIX: csrf_exempt hataya, csrf_protect import rakha (reference ke liye)
from django.conf import settings
from django.core import serializers
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint

from .models import (
    Goat, BreedingRecord, HealthRecord, MilkProduction,
    FeedInventory, FeedConsumption, Sale, Expense, WeightRecord,
    PerformanceEvaluation, MarketPrice, WeatherRecord, FarmEvent,
    BreedingPlan, CustomReminder, AdditionalIncome, Task,
    Customer, Credit, Notification, Insurance, MortalityRecord,
    VaccinationSchedule, BudgetPlanning, ActivityLog, VetVisit,
    Document, PhotoGallery,
)
from django.db.models import Sum, Count, Q


# ─────────────────────────────────────────────
#  STYLING CONSTANTS
# ─────────────────────────────────────────────

HDR_GREEN  = PatternFill("solid", start_color="1F6B35", end_color="1F6B35")
HDR_BLUE   = PatternFill("solid", start_color="2F4F8F", end_color="2F4F8F")
HDR_AMBER  = PatternFill("solid", start_color="B8860B", end_color="B8860B")
HDR_RED    = PatternFill("solid", start_color="8B1A1A", end_color="8B1A1A")
ROW_EVEN   = PatternFill("solid", start_color="F4FAF6", end_color="F4FAF6")
ROW_ODD    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
ACCENT     = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")

HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TITLE_FONT = Font(bold=True, color="1F6B35", name="Arial", size=14)
BODY_FONT  = Font(name="Arial", size=9)
MONEY_FONT = Font(name="Arial", size=9, color="1A5C2A")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
MEDIUM_BORDER = Border(
    left=Side(style='medium', color='1F6B35'),
    right=Side(style='medium', color='1F6B35'),
    top=Side(style='medium', color='1F6B35'),
    bottom=Side(style='medium', color='1F6B35'),
)

INR_FORMAT  = '₹#,##0.00'
DATE_FORMAT = 'DD-MMM-YYYY'
NUM_FORMAT  = '#,##0.00'


# ─────────────────────────────────────────────
#  HELPER FUNCTIONS
# ─────────────────────────────────────────────

def _hdr(ws, headers, row=1, fill=HDR_GREEN):
    """Write styled header row."""
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill  = fill
        c.font  = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def _body(ws, data_rows, start_row=2):
    """Write body rows with alternating colors."""
    for r_idx, row_data in enumerate(data_rows, start_row):
        fill = ROW_EVEN if r_idx % 2 == 0 else ROW_ODD
        for c_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill   = fill
            c.font   = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical='center', wrap_text=False)
        ws.row_dimensions[r_idx].height = 16


def _auto_width(ws, min_w=8, max_w=40):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 3, min_w), max_w)


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _title_row(ws, title, col_span, row=1):
    """Full-width title above headers."""
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    c.fill = PatternFill("solid", start_color="0D4A20", end_color="0D4A20")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26


def _safe_date(d):
    return d.strftime('%d-%m-%Y') if d else ''


def _safe_dt(dt):
    return dt.strftime('%d-%m-%Y %H:%M') if dt else ''


# ─────────────────────────────────────────────
#  SHEET BUILDERS
# ─────────────────────────────────────────────

def _sheet_goats(wb):
    ws = wb.create_sheet("🐐 Goats")
    hdrs = ['Tag No.', 'Name', 'Breed', 'Gender', 'Color', 'Date of Birth',
            'Age (Yrs)', 'Weight (kg)', 'Purchase Date', 'Purchase Price (₹)',
            'Status', 'Mother', 'Father', 'Created At']
    _hdr(ws, hdrs)
    rows = []
    for g in Goat.objects.select_related('mother','father').order_by('tag_number'):
        rows.append([
            g.tag_number, g.name,
            dict(Goat.BREED_CHOICES).get(g.breed, g.breed),
            'Male' if g.gender == 'M' else 'Female',
            g.color, _safe_date(g.date_of_birth), g.get_age_years(),
            g.weight, _safe_date(g.purchase_date), g.purchase_price,
            dict(Goat.STATUS_CHOICES).get(g.status, g.status),
            g.mother.name if g.mother else '',
            g.father.name if g.father else '',
            _safe_dt(g.created_at),
        ])
    _body(ws, rows)
    # Money column formatting
    for row in ws.iter_rows(min_row=3, min_col=10, max_col=10):
        for c in row:
            c.number_format = INR_FORMAT
            c.font = MONEY_FONT
    _title_row(ws, f"🐐  GOAT INVENTORY  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_breeding(wb):
    ws = wb.create_sheet("🤝 Breeding")
    hdrs = ['Mother', 'Father', 'Breeding Date', 'Expected Delivery',
            'Actual Delivery', 'Status', 'No. of Kids', 'Notes', 'Created At']
    _hdr(ws, hdrs, fill=HDR_BLUE)
    rows = []
    for r in BreedingRecord.objects.select_related('mother','father').order_by('-breeding_date'):
        rows.append([
            r.mother.name, r.father.name,
            _safe_date(r.breeding_date), _safe_date(r.expected_delivery_date),
            _safe_date(r.actual_delivery_date),
            dict(BreedingRecord.STATUS_CHOICES).get(r.status, r.status),
            r.number_of_kids or '', r.notes, _safe_dt(r.created_at),
        ])
    _body(ws, rows)
    _title_row(ws, f"🤝  BREEDING RECORDS  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_health(wb):
    ws = wb.create_sheet("🏥 Health")
    hdrs = ['Goat', 'Type', 'Date', 'Description', 'Medicine',
            'Dosage', 'Cost (₹)', 'Veterinarian', 'Next Due', 'Created At']
    _hdr(ws, hdrs, fill=HDR_RED)
    rows = []
    total_cost = 0
    for r in HealthRecord.objects.select_related('goat').order_by('-date'):
        rows.append([
            r.goat.name,
            dict(HealthRecord.RECORD_TYPE_CHOICES).get(r.record_type, r.record_type),
            _safe_date(r.date), r.description, r.medicine_used,
            r.dosage, r.cost, r.veterinarian,
            _safe_date(r.next_due_date), _safe_dt(r.created_at),
        ])
        total_cost += r.cost
    _body(ws, rows)
    # Total row
    total_r = len(rows) + 3
    ws.cell(total_r, 1, "TOTAL HEALTH COST").font = Font(bold=True, name="Arial", size=9)
    ws.cell(total_r, 7, total_cost).number_format = INR_FORMAT
    ws.cell(total_r, 7).font = Font(bold=True, color="8B1A1A", name="Arial", size=9)
    # Format cost column
    for row in ws.iter_rows(min_row=3, min_col=7, max_col=7):
        for c in row:
            c.number_format = INR_FORMAT
    _title_row(ws, f"🏥  HEALTH RECORDS  —  Total Cost: ₹{total_cost:,.0f}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_milk(wb):
    ws = wb.create_sheet("🥛 Milk")
    hdrs = ['Goat', 'Date', 'Session', 'Quantity (L)', 'Fat %', 'Created At']
    _hdr(ws, hdrs, fill=HDR_GREEN)
    rows = []
    total_milk = 0
    for r in MilkProduction.objects.select_related('goat').order_by('-date'):
        rows.append([
            r.goat.name, _safe_date(r.date),
            'Morning' if r.session == 'M' else 'Evening',
            r.quantity, r.fat_percentage or '', _safe_dt(r.created_at),
        ])
        total_milk += r.quantity
    _body(ws, rows)
    total_r = len(rows) + 3
    ws.cell(total_r, 1, f"TOTAL: {total_milk:.1f} L").font = Font(bold=True, name="Arial", size=9)
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=4):
        for c in row:
            c.number_format = NUM_FORMAT
    _title_row(ws, f"🥛  MILK PRODUCTION  —  Total: {total_milk:.1f} Litres", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_sales(wb):
    ws = wb.create_sheet("💰 Sales")
    hdrs = ['Type', 'Goat', 'Date', 'Qty', 'Unit', 'Price/Unit (₹)',
            'Total Amount (₹)', 'Buyer', 'Contact', 'Payment', 'Created At']
    _hdr(ws, hdrs, fill=HDR_GREEN)
    rows = []
    total_rev = 0
    type_map = {'G':'Goat','M':'Milk','MN':'Manure','O':'Other'}
    pay_map  = {'P':'Paid','UP':'Unpaid','PA':'Partial'}
    for s in Sale.objects.select_related('goat').order_by('-date'):
        rows.append([
            type_map.get(s.sale_type, s.sale_type),
            s.goat.name if s.goat else '',
            _safe_date(s.date), s.quantity, s.unit,
            s.price_per_unit, s.total_amount,
            s.buyer_name, s.buyer_contact,
            pay_map.get(s.payment_status, s.payment_status),
            _safe_dt(s.created_at),
        ])
        total_rev += s.total_amount
    _body(ws, rows)
    total_r = len(rows) + 3
    ws.cell(total_r, 1, "TOTAL REVENUE").font = Font(bold=True, name="Arial", size=9)
    ws.cell(total_r, 7, total_rev).number_format = INR_FORMAT
    ws.cell(total_r, 7).font = Font(bold=True, color="1A5C2A", name="Arial", size=9)
    for row in ws.iter_rows(min_row=3, min_col=6, max_col=7):
        for c in row:
            c.number_format = INR_FORMAT
    _title_row(ws, f"💰  SALES  —  Total Revenue: ₹{total_rev:,.0f}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_expenses(wb):
    ws = wb.create_sheet("💸 Expenses")
    hdrs = ['Date', 'Type', 'Description', 'Amount (₹)', 'Paid To', 'Method', 'Created At']
    _hdr(ws, hdrs, fill=HDR_AMBER)
    rows = []
    total_exp = 0
    type_map = {'F':'Feed','M':'Medicine','V':'Veterinary','R':'Repairs',
                'U':'Utilities','L':'Labour','O':'Other'}
    for e in Expense.objects.order_by('-date'):
        rows.append([
            _safe_date(e.date),
            type_map.get(e.expense_type, e.expense_type),
            e.description, e.amount, e.paid_to, e.payment_method,
            _safe_dt(e.created_at),
        ])
        total_exp += e.amount
    _body(ws, rows)
    total_r = len(rows) + 3
    ws.cell(total_r, 1, "TOTAL EXPENSES").font = Font(bold=True, name="Arial", size=9)
    ws.cell(total_r, 4, total_exp).number_format = INR_FORMAT
    ws.cell(total_r, 4).font = Font(bold=True, color="8B1A1A", name="Arial", size=9)
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=4):
        for c in row:
            c.number_format = INR_FORMAT
    _title_row(ws, f"💸  EXPENSES  —  Total: ₹{total_exp:,.0f}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_weight(wb):
    ws = wb.create_sheet("⚖️ Weight")
    hdrs = ['Goat', 'Tag No.', 'Date', 'Weight (kg)', 'Recorded At']
    _hdr(ws, hdrs, fill=HDR_BLUE)
    rows = []
    for r in WeightRecord.objects.select_related('goat').order_by('-date'):
        rows.append([
            r.goat.name, r.goat.tag_number, _safe_date(r.date),
            r.weight, _safe_dt(r.created_at),
        ])
    _body(ws, rows)
    _title_row(ws, f"⚖️  WEIGHT RECORDS  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_tasks(wb):
    ws = wb.create_sheet("📋 Tasks")
    hdrs = ['Title', 'Priority', 'Status', 'Due Date', 'Assigned To', 'Description', 'Created At']
    _hdr(ws, hdrs, fill=HDR_BLUE)
    rows = []
    priority_map = {'H':'High','M':'Medium','L':'Low'}
    status_map   = {'P':'Pending','IP':'In Progress','C':'Completed'}
    for t in Task.objects.order_by('due_date'):
        rows.append([
            t.title,
            priority_map.get(t.priority, t.priority),
            status_map.get(t.status, t.status),
            _safe_date(t.due_date), t.assigned_to, t.description,
            _safe_dt(t.created_at),
        ])
    _body(ws, rows)
    _title_row(ws, f"📋  TASKS  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_vaccination(wb):
    ws = wb.create_sheet("💉 Vaccination")
    hdrs = ['Goat', 'Tag No.', 'Vaccine', 'Due Date', 'Completed', 'Completion Date', 'Created At']
    _hdr(ws, hdrs, fill=HDR_RED)
    rows = []
    for r in VaccinationSchedule.objects.select_related('goat').order_by('due_date'):
        rows.append([
            r.goat.name, r.goat.tag_number, r.vaccine_name,
            _safe_date(r.due_date), 'Yes' if r.completed else 'No',
            _safe_date(r.completion_date), _safe_dt(r.created_at),
        ])
    _body(ws, rows)
    _title_row(ws, f"💉  VACCINATION SCHEDULE  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_insurance(wb):
    ws = wb.create_sheet("🛡️ Insurance")
    hdrs = ['Goat', 'Provider', 'Policy No.', 'Coverage (₹)', 'Premium (₹)',
            'Start Date', 'End Date', 'Created At']
    _hdr(ws, hdrs, fill=HDR_BLUE)
    rows = []
    for r in Insurance.objects.select_related('goat').order_by('end_date'):
        rows.append([
            r.goat.name, r.provider, r.policy_number,
            r.coverage_amount, r.premium,
            _safe_date(r.start_date), _safe_date(r.end_date),
            _safe_dt(r.created_at),
        ])
    _body(ws, rows)
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=5):
        for c in row:
            c.number_format = INR_FORMAT
    _title_row(ws, f"🛡️  INSURANCE  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_customers(wb):
    ws = wb.create_sheet("👥 Customers")
    hdrs = ['Name', 'Contact', 'Email', 'Address', 'Created At']
    _hdr(ws, hdrs, fill=HDR_GREEN)
    rows = []
    for c in Customer.objects.order_by('name'):
        rows.append([c.name, c.contact, c.email, c.address, _safe_dt(c.created_at)])
    _body(ws, rows)
    _title_row(ws, f"👥  CUSTOMERS  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_feed(wb):
    ws = wb.create_sheet("🌾 Feed")
    hdrs = ['Feed Name', 'Type', 'Qty (kg)', 'Unit Price (₹)', 'Total Cost (₹)',
            'Purchase Date', 'Supplier', 'Created At']
    _hdr(ws, hdrs, fill=HDR_GREEN)
    type_map = {'G':'Green','D':'Dry','C':'Concentrate','H':'Hay','S':'Supplement'}
    rows = []
    for f in FeedInventory.objects.order_by('-purchase_date'):
        rows.append([
            f.feed_name, type_map.get(f.feed_type, f.feed_type),
            f.quantity, f.unit_price, f.total_cost(),
            _safe_date(f.purchase_date), f.supplier, _safe_dt(f.created_at),
        ])
    _body(ws, rows)
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=5):
        for c in row:
            c.number_format = INR_FORMAT
    _title_row(ws, f"🌾  FEED INVENTORY  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_mortality(wb):
    ws = wb.create_sheet("💀 Mortality")
    hdrs = ['Goat', 'Tag No.', 'Death Date', 'Cause', 'Age at Death (months)', 'Weight at Death (kg)', 'Created At']
    _hdr(ws, hdrs, fill=HDR_RED)
    rows = []
    for r in MortalityRecord.objects.select_related('goat').order_by('-death_date'):
        rows.append([
            r.goat.name, r.goat.tag_number, _safe_date(r.death_date),
            r.cause, r.age_at_death, r.weight_at_death, _safe_dt(r.created_at),
        ])
    _body(ws, rows)
    _title_row(ws, f"💀  MORTALITY RECORDS  —  Total: {len(rows)}", len(hdrs))
    _auto_width(ws)
    _freeze(ws, "A3")
    return ws


def _sheet_summary(wb, timestamp):
    """Professional summary / dashboard sheet — first sheet."""
    ws = wb.create_sheet("📊 Summary", 0)

    # ── Stats ──
    total_goats    = Goat.objects.count()
    active_goats   = Goat.objects.filter(status='A').count()
    pregnant_goats = Goat.objects.filter(status='P').count()
    sold_goats     = Goat.objects.filter(status='S').count()
    dead_goats     = Goat.objects.filter(status='D').count()
    total_milk     = MilkProduction.objects.aggregate(t=Sum('quantity'))['t'] or 0
    total_revenue  = Sale.objects.aggregate(t=Sum('total_amount'))['t'] or 0
    total_expenses = Expense.objects.aggregate(t=Sum('amount'))['t'] or 0
    net_profit     = total_revenue - total_expenses
    health_cost    = HealthRecord.objects.aggregate(t=Sum('cost'))['t'] or 0

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    # Title
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = "🐐  GOAT FARM MANAGEMENT — DATA EXPORT SUMMARY"
    t.font  = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    t.fill  = PatternFill("solid", start_color="0D4A20", end_color="0D4A20")
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Export info
    ws['A2'] = "Export Date:"
    ws['B2'] = timestamp
    ws['A2'].font = Font(bold=True, name="Arial", size=9, color="555555")
    ws['B2'].font = Font(name="Arial", size=9)

    def section(row, label, fill):
        ws.merge_cells(f'A{row}:D{row}')
        c = ws.cell(row, 1, label)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20

    def stat_row(row, label, value, fmt=None, good=None):
        c1 = ws.cell(row, 1, label)
        c2 = ws.cell(row, 2, value)
        c1.font = Font(name="Arial", size=9)
        c2.font = Font(name="Arial", size=10, bold=True,
                       color=("1A5C2A" if good else ("8B1A1A" if good is False else "222222")))
        c1.fill = c2.fill = ROW_EVEN if row % 2 == 0 else ROW_ODD
        c1.border = c2.border = THIN_BORDER
        c1.alignment = Alignment(vertical='center')
        c2.alignment = Alignment(horizontal='right', vertical='center')
        if fmt:
            c2.number_format = fmt
        ws.row_dimensions[row].height = 16

    r = 4
    section(r, "  🐐  GOAT STATISTICS", HDR_GREEN); r+=1
    stat_row(r, "Total Goats",     total_goats);    r+=1
    stat_row(r, "Active (A)",      active_goats, good=True);   r+=1
    stat_row(r, "Pregnant (P)",    pregnant_goats); r+=1
    stat_row(r, "Sold (S)",        sold_goats);     r+=1
    stat_row(r, "Dead (D)",        dead_goats, good=(dead_goats==0)); r+=1

    r+=1
    section(r, "  💰  FINANCIALS (All Time)", HDR_GREEN); r+=1
    stat_row(r, "Total Revenue (₹)",  total_revenue,  INR_FORMAT, good=True);   r+=1
    stat_row(r, "Total Expenses (₹)", total_expenses, INR_FORMAT, good=False);  r+=1
    stat_row(r, "Net Profit (₹)",     net_profit,     INR_FORMAT, good=(net_profit>=0)); r+=1
    stat_row(r, "Health Cost (₹)",    health_cost,    INR_FORMAT); r+=1
    stat_row(r, "Milk Produced (L)",  round(total_milk, 1), NUM_FORMAT, good=True); r+=1

    r+=1
    section(r, "  📋  RECORD COUNTS", HDR_BLUE); r+=1
    stat_row(r, "Breeding Records",   BreedingRecord.objects.count()); r+=1
    stat_row(r, "Health Records",     HealthRecord.objects.count());   r+=1
    stat_row(r, "Milk Records",       MilkProduction.objects.count()); r+=1
    stat_row(r, "Sales Records",      Sale.objects.count());           r+=1
    stat_row(r, "Expense Records",    Expense.objects.count());        r+=1
    stat_row(r, "Weight Records",     WeightRecord.objects.count());   r+=1
    stat_row(r, "Tasks",              Task.objects.count());           r+=1
    stat_row(r, "Customers",          Customer.objects.count());       r+=1
    stat_row(r, "Vaccination Schedules", VaccinationSchedule.objects.count()); r+=1
    stat_row(r, "Insurance Policies", Insurance.objects.count());     r+=1
    stat_row(r, "Mortality Records",  MortalityRecord.objects.count()); r+=1
    stat_row(r, "Feed Items",         FeedInventory.objects.count()); r+=1

    return ws


# ─────────────────────────────────────────────
#  JSON BACKUP / RESTORE
# ─────────────────────────────────────────────

ALL_MODELS = [
    ('goats',        Goat),
    ('breeding',     BreedingRecord),
    ('health',       HealthRecord),
    ('milk',         MilkProduction),
    ('feed_inventory', FeedInventory),
    ('feed_consumption', FeedConsumption),
    ('sales',        Sale),
    ('expenses',     Expense),
    ('weight',       WeightRecord),
    ('performance',  PerformanceEvaluation),
    ('market_prices', MarketPrice),
    ('weather_records', WeatherRecord),
    ('farm_events',  FarmEvent),
    ('breeding_plans', BreedingPlan),
    ('reminders',    CustomReminder),
    ('additional_income', AdditionalIncome),
    ('tasks',        Task),
    ('customers',    Customer),
    ('credits',      Credit),
    ('notifications', Notification),
    ('insurance',    Insurance),
    ('mortality',    MortalityRecord),
    ('vaccination_schedule', VaccinationSchedule),
    ('budget_planning', BudgetPlanning),
    ('activity_log', ActivityLog),
    ('vet_visits',   VetVisit),
    ('documents',    Document),
    ('photo_gallery', PhotoGallery),
]


def _build_json_backup():
    """Sab models ka dict return karo — JSON-serializable."""
    backup = {
        'meta': {
            'version': '5.4',
            'created_at': datetime.now().isoformat(),
            'total_models': len(ALL_MODELS),
        },
        'data': {}
    }
    for key, Model in ALL_MODELS:
        try:
            qs = Model.objects.all()
            backup['data'][key] = json.loads(
                serializers.serialize('json', qs)
            )
            backup['meta'][f'{key}_count'] = qs.count()
        except Exception as e:
            backup['data'][key] = []
            backup['meta'][f'{key}_error'] = str(e)

    return backup


# ─────────────────────────────────────────────
#  VIEWS
# ─────────────────────────────────────────────

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def backup_page(request):
    """Backup & Download page render karo."""
    from django.shortcuts import render

    # Stats for the page
    stats = {
        'goats':    Goat.objects.count(),
        'sales':    Sale.objects.count(),
        'health':   HealthRecord.objects.count(),
        'milk':     MilkProduction.objects.count(),
        'expenses': Expense.objects.count(),
        'tasks':    Task.objects.count(),
        'customers': Customer.objects.count(),
        'total_revenue': Sale.objects.aggregate(t=Sum('total_amount'))['t'] or 0,
        'total_expenses': Expense.objects.aggregate(t=Sum('amount'))['t'] or 0,
    }
    return render(request, 'farm/backup.html', {'stats': stats})


# ── JSON BACKUP ───────────────────────────────

@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_json_backup(request):
    """Complete JSON backup download — sab models."""
    backup = _build_json_backup()
    filename = f"goat_farm_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    response = HttpResponse(
        json.dumps(backup, indent=2, ensure_ascii=False, default=str),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_zip_backup(request):
    """JSON + all Excel sheets in a single ZIP."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # JSON backup
        backup = _build_json_backup()
        zf.writestr(
            f"backup_{timestamp}/data_backup.json",
            json.dumps(backup, indent=2, ensure_ascii=False, default=str)
        )
        # Excel — complete workbook
        wb = _build_complete_workbook(timestamp)
        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        zf.writestr(
            f"backup_{timestamp}/farm_data_{timestamp}.xlsx",
            excel_buf.getvalue()
        )
        # README
        readme = (
            f"GOAT FARM BACKUP — {timestamp}\n"
            f"{'='*40}\n\n"
            f"Files:\n"
            f"  data_backup.json  — Full JSON backup (restore ke liye)\n"
            f"  farm_data.xlsx    — Excel export (viewing ke liye)\n\n"
            f"Restore karne ke liye:\n"
            f"  POST /backup/restore/  mein JSON file upload karein\n"
        )
        zf.writestr(f"backup_{timestamp}/README.txt", readme)

    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="farm_backup_{timestamp}.zip"'
    return response


@login_required(login_url='/login/')
@require_http_methods(["POST"])
def restore_json_backup(request):
    """
    JSON backup se data restore karo.
    POST body: multipart/form-data — 'backup_file' field mein .json file.
    WARNING: Ye existing data DELETE karta hai phir restore karta hai.
    SECURITY: Sirf staff users hi restore kar sakte hain.
    """
    # FIX: @csrf_exempt hataya — security risk tha. Frontend X-CSRFToken header bheje.
    # FIX: Staff-only check add kiya — yeh destructive operation hai
    if not request.user.is_staff:
        return JsonResponse({'error': 'Sirf admin/staff users hi backup restore kar sakte hain.'}, status=403)
    if 'backup_file' not in request.FILES:
        return JsonResponse({'error': 'backup_file field missing'}, status=400)

    f = request.FILES['backup_file']
    try:
        raw = f.read().decode('utf-8')
        backup = json.loads(raw)
    except Exception as e:
        return JsonResponse({'error': f'Invalid JSON: {e}'}, status=400)

    if 'data' not in backup:
        return JsonResponse({'error': "File mein 'data' key nahi mili — valid backup nahi hai"}, status=400)

    results = {}
    errors  = {}

    with transaction.atomic():
        # Delete in reverse order to avoid FK violations
        for key, Model in reversed(ALL_MODELS):
            try:
                deleted, _ = Model.objects.all().delete()
                results[key] = {'deleted': deleted}
            except Exception as e:
                errors[key] = {'delete_error': str(e)}

        # Restore
        for key, Model in ALL_MODELS:
            if key not in backup['data']:
                continue
            try:
                objects = list(serializers.deserialize('json', json.dumps(backup['data'][key])))
                for obj in objects:
                    obj.save()
                results[key]['restored'] = len(objects)
            except Exception as e:
                errors[key] = {'restore_error': str(e)}

    return JsonResponse({
        'success': True,
        'restored_at': datetime.now().isoformat(),
        'results': results,
        'errors': errors,
    })


# ── EXCEL DOWNLOADS ───────────────────────────

def _build_complete_workbook(timestamp=None):
    """Sab sheets wala workbook banao."""
    if not timestamp:
        timestamp = datetime.now().strftime('%d-%m-%Y %H:%M')
    wb = Workbook()
    wb.remove(wb.active)  # blank sheet hata do

    _sheet_summary(wb, timestamp)
    _sheet_goats(wb)
    _sheet_breeding(wb)
    _sheet_health(wb)
    _sheet_milk(wb)
    _sheet_sales(wb)
    _sheet_expenses(wb)
    _sheet_weight(wb)
    _sheet_tasks(wb)
    _sheet_vaccination(wb)
    _sheet_insurance(wb)
    _sheet_customers(wb)
    _sheet_feed(wb)
    _sheet_mortality(wb)

    # Workbook-level styling
    wb.properties.creator = "Goat Farm Management System v5.4"
    wb.properties.title   = "Goat Farm — Complete Data Export"

    return wb


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def download_complete_excel(request):
    """Sab data ek Excel mein — sab sheets ke saath."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    wb = _build_complete_workbook(datetime.now().strftime('%d-%m-%Y %H:%M:%S'))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="farm_complete_{timestamp}.xlsx"'
    return response


def _single_sheet_response(builder_fn, filename_prefix):
    """Helper: single sheet download."""
    @login_required(login_url='/login/')
    @require_http_methods(["GET"])
    def view(request):
        wb = Workbook()
        wb.remove(wb.active)
        builder_fn(wb)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timestamp}.xlsx"'
        return resp
    return view


# Individual sheet downloads
download_excel_goats       = _single_sheet_response(_sheet_goats,      'goats')
download_excel_breeding    = _single_sheet_response(_sheet_breeding,   'breeding')
download_excel_health      = _single_sheet_response(_sheet_health,     'health')
download_excel_milk        = _single_sheet_response(_sheet_milk,       'milk')
download_excel_sales       = _single_sheet_response(_sheet_sales,      'sales')
download_excel_expenses    = _single_sheet_response(_sheet_expenses,   'expenses')
download_excel_weight      = _single_sheet_response(_sheet_weight,     'weight')
download_excel_tasks       = _single_sheet_response(_sheet_tasks,      'tasks')
download_excel_vaccination = _single_sheet_response(_sheet_vaccination,'vaccination')
download_excel_insurance   = _single_sheet_response(_sheet_insurance,  'insurance')
download_excel_customers   = _single_sheet_response(_sheet_customers,  'customers')
download_excel_feed        = _single_sheet_response(_sheet_feed,       'feed')
download_excel_mortality   = _single_sheet_response(_sheet_mortality,  'mortality')


@login_required(login_url='/login/')
@require_http_methods(["GET"])
def backup_stats_api(request):
    """Backup page ke liye live stats — JSON."""
    return JsonResponse({
        'goats':    Goat.objects.count(),
        'sales':    Sale.objects.count(),
        'health':   HealthRecord.objects.count(),
        'milk':     MilkProduction.objects.count(),
        'expenses': Expense.objects.count(),
        'weight':   WeightRecord.objects.count(),
        'tasks':    Task.objects.count(),
        'customers': Customer.objects.count(),
        'breeding': BreedingRecord.objects.count(),
        'vaccination': VaccinationSchedule.objects.count(),
        'insurance': Insurance.objects.count(),
        'mortality': MortalityRecord.objects.count(),
        'total_revenue':  Sale.objects.aggregate(t=Sum('total_amount'))['t'] or 0,
        'total_expenses': Expense.objects.aggregate(t=Sum('amount'))['t'] or 0,
        'total_milk': MilkProduction.objects.aggregate(t=Sum('quantity'))['t'] or 0,
        'generated_at': datetime.now().isoformat(),
    })
